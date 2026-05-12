"""POF Dashboard generator.

Reads the POF leader-list + 1-1 tracking spreadsheet and regenerates index.html.

Sources, in order of preference:
  1. Google Sheet (via service-account key in env $GOOGLE_SERVICE_ACCOUNT_JSON)
  2. Local xlsx at ~/Desktop/POF Leader LIST.xlsx (fallback for local builds)

Outputs:
  - index.html (rendered from scripts/template.html)
  - data/data.json (this build's stats)
  - data/data-previous.json (prior build's stats — used for week-over-week deltas)
"""
from __future__ import annotations
import datetime as dt
import json
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

# ── CONFIG ──────────────────────────────────────────────────────────────────
TODAY = dt.date.today()
ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
TEMPLATE = ROOT / "scripts" / "template.html"
OUTPUT_HTML = ROOT / "index.html"
# Search common locations for the xlsx (Liz moves it around)
_XLSX_CANDIDATES = [
    Path.home() / "Desktop" / "POF Leader LIST.xlsx",
    Path.home() / "Desktop" / "eb" / "POF Leader LIST.xlsx",
    Path.home() / "Downloads" / "POF Leader LIST.xlsx",
]
LOCAL_XLSX = next((p for p in _XLSX_CANDIDATES if p.exists()), _XLSX_CANDIDATES[0])
SHEET_ID = "1v2b7OMQ3Hvz9LwMroYw23CHIn3NsBZhFxDezy-QzcXQ"

# Header aliases — when fellows customize their tabs, map their headers back to canonical names.
HEADER_ALIASES = {
    "first_name": ["first name", "firstname", "first"],
    "last_name":  ["last name", "lastname", "last"],
    "district":   ["district or org", "district", "org", "organization"],
    "school":     ["school", "school name"],
    "level":      ["leadership level", "level"],
    "self_interest": ["self-interest", "self interest", "what is their self-interest?",
                     "what is their self interest?", "self-interest?"],
    "next_step":  ["next risk / action / invitation", "next steps", "next step"],
    "inv_status": ["invitation status", "status"],
    "date":       ["date of 1-1", "date of 1:1", "last 1:1", "last 1-1", "1-1 date", "date"],
    "notes":      ["notes from 1-1", "notes from 1:1", "notes"],
    "source":     ["how did you find them?", "how did you find them", "source", "how found"],
    "reflection": ["reflections on your 1-1 leadership?", "reflection"],
}

# ── HELPERS ─────────────────────────────────────────────────────────────────

def norm(s):
    if s is None: return ""
    return str(s).strip().lower().rstrip("?").strip()

def build_header_map(ws):
    """Return {canonical_key: column_index} from the worksheet's first row."""
    # Pre-normalize alias dictionary (strip ? trailing whitespace etc.) so it matches normed cells
    aliases_normed = {k: [norm(a) for a in v] for k, v in HEADER_ALIASES.items()}
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if not v: continue
        n = norm(v)
        for canonical, aliases in aliases_normed.items():
            if n in aliases:
                headers[canonical] = c
                break
    return headers

def initial(name):
    """First name + last initial for an entry. 'Jill Leach' → 'Jill L.'"""
    if not name: return ""
    parts = str(name).strip().split()
    if not parts: return ""
    if len(parts) == 1: return parts[0]
    return f"{parts[0]} {parts[-1][0]}."

def parse_date(d):
    """Parse any reasonable date string. Returns datetime.date or None."""
    if d is None: return None
    if isinstance(d, dt.datetime): return d.date()
    if isinstance(d, dt.date): return d
    s = str(d).strip()
    if not s: return None
    sl = s.lower()
    if any(t in sl for t in [
        "reschedule", "pending", "requested", "tbd", "n/a",
        "follow up needed", "left message", "not contacted",
    ]):
        return None
    m = re.search(r"(\d{1,2})[./-](\d{1,2})(?:[./-](\d{2,4}))?", s)
    if not m: return None
    mo, dd = int(m.group(1)), int(m.group(2))
    y = m.group(3)
    if y:
        y = int(y)
        if y < 100: y += 2000
    else:
        y = TODAY.year
    try: return dt.date(y, mo, dd)
    except Exception: return None

def truncate(text, max_words=25):
    if not text: return ""
    words = str(text).strip().split()
    if len(words) <= max_words: return " ".join(words)
    return " ".join(words[:max_words]) + "…"

# ── DATA LOADING ────────────────────────────────────────────────────────────

def load_workbook_local():
    import openpyxl
    return openpyxl.load_workbook(LOCAL_XLSX, data_only=True)

def load_workbook_sheets():
    """Read every tab into an openpyxl-like object via gspread."""
    key_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not key_json: return None
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("→ gspread not installed; falling back to local xlsx", file=sys.stderr)
        return None
    info = json.loads(key_json)
    creds = Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"])
    gc = gspread.authorize(creds)
    return gc.open_by_key(SHEET_ID)

def fellow_names(workbook):
    """Return sorted list of fellow names from tab names like 'David - Leader list'."""
    if hasattr(workbook, "sheetnames"):  # openpyxl
        sheets = workbook.sheetnames
    else:  # gspread spreadsheet
        sheets = [w.title for w in workbook.worksheets()]
    fellows = set()
    for s in sheets:
        if " - " not in s: continue
        name, kind = s.split(" - ", 1)
        if name == "TEMPLATE": continue
        fellows.add(name)
    return sorted(fellows)

class Sheet:
    """Uniform interface over openpyxl Worksheet or gspread Worksheet."""
    def __init__(self, ws, kind):
        self.ws = ws
        self.kind = kind  # 'xlsx' or 'gsheet'
        if kind == "xlsx":
            self.max_row = ws.max_row
            self.max_column = ws.max_column
        else:
            # gspread: get_all_values gives us a 2-D list of strings
            self.values = ws.get_all_values()
            self.max_row = len(self.values)
            self.max_column = max((len(r) for r in self.values), default=0)
    def cell(self, row, column):
        if self.kind == "xlsx":
            return self.ws.cell(row=row, column=column).value
        # gspread returns strings only; we re-parse dates downstream
        try: return self.values[row-1][column-1]
        except IndexError: return None

def get_sheet(workbook, sheetname):
    if hasattr(workbook, "sheetnames"):  # openpyxl
        if sheetname not in workbook.sheetnames: return None
        return Sheet(workbook[sheetname], "xlsx")
    # gspread
    try: return Sheet(workbook.worksheet(sheetname), "gsheet")
    except Exception: return None

# ── AGGREGATION ─────────────────────────────────────────────────────────────

def aggregate(workbook):
    """Build the full data model for one build."""
    fellows = fellow_names(workbook)
    out = {
        "as_of": TODAY.isoformat(),
        "fellows": {},
        "totals": {},
        "all_completed": [],   # for voices section
        "schools_to_fellows": defaultdict(list),
        "source_buckets": Counter(),
        "warnings": [],
    }

    for f in fellows:
        info = {
            "name": f,
            "leaders": [],
            "prospects": [],
            "completed": [],
            "scheduled": [],
            "districts": set(),
            "schools": set(),
            "leader_levels": Counter(),
            "inv_statuses": Counter(),
            "source_buckets": Counter(),
            "chains": [],  # (hub_name, [referred_names])
        }

        # Leader list
        ll = get_sheet(workbook, f"{f} - Leader list")
        if ll:
            h = build_header_map(ll.ws if ll.kind == "xlsx" else _GspreadHeaderShim(ll.values))
            if "first_name" not in h:
                out["warnings"].append(f"{f}: no First Name column in Leader list")
            else:
                for r in range(2, ll.max_row + 1):
                    fn = ll.cell(r, h["first_name"])
                    ln = ll.cell(r, h.get("last_name", 0)) if "last_name" in h else None
                    if not (fn or ln): continue
                    leader = {
                        "name": initial(f"{(fn or '').strip()} {(ln or '').strip()}"),
                        "district": (ll.cell(r, h["district"]) if "district" in h else "") or "",
                        "school":   (ll.cell(r, h["school"])   if "school"   in h else "") or "",
                        "level":    (ll.cell(r, h["level"])    if "level"    in h else "") or "",
                        "self_interest": truncate(ll.cell(r, h["self_interest"]) if "self_interest" in h else ""),
                        "next_step": truncate(ll.cell(r, h["next_step"]) if "next_step" in h else ""),
                    }
                    info["leaders"].append(leader)
                    if leader["district"]: info["districts"].add(str(leader["district"]).strip())
                    if leader["school"]:
                        s = str(leader["school"]).strip()
                        info["schools"].add(s)
                        out["schools_to_fellows"][s.lower()].append(f)
                    if leader["level"]: info["leader_levels"][str(leader["level"]).strip()] += 1

        # 1-1 Tracking
        tt = get_sheet(workbook, f"{f} - 1-1 Tracking")
        if tt:
            h = build_header_map(tt.ws if tt.kind == "xlsx" else _GspreadHeaderShim(tt.values))
            if "first_name" not in h:
                out["warnings"].append(f"{f}: no First Name column in 1-1 Tracking")
            else:
                for r in range(2, tt.max_row + 1):
                    fn = tt.cell(r, h["first_name"])
                    ln = tt.cell(r, h.get("last_name", 0)) if "last_name" in h else None
                    if not (fn or ln): continue
                    full = f"{(fn or '').strip()} {(ln or '').strip()}".strip()
                    raw_date = tt.cell(r, h["date"]) if "date" in h else None
                    d = parse_date(raw_date)
                    notes = tt.cell(r, h["notes"]) if "notes" in h else None
                    si = tt.cell(r, h["self_interest"]) if "self_interest" in h else None
                    src = tt.cell(r, h["source"]) if "source" in h else None
                    school = (tt.cell(r, h["school"]) if "school" in h else "") or ""
                    inv = (tt.cell(r, h["inv_status"]) if "inv_status" in h else "") or ""
                    entry = {
                        "name": initial(full),
                        "full_name_lower": full.lower(),
                        "school": str(school).strip(),
                        "date": d.isoformat() if d else None,
                        "raw_date": str(raw_date).strip() if raw_date else "",
                        "has_notes": bool(notes and str(notes).strip()),
                        "has_si": bool(si and str(si).strip()),
                        "self_interest_text": truncate(si, 30),
                        "next_step_text": truncate(tt.cell(r, h["next_step"]) if "next_step" in h else "", 20),
                        "source": str(src).strip() if src else "",
                        "inv_status": str(inv).strip(),
                    }
                    info["prospects"].append(entry)
                    if d and d <= TODAY: info["completed"].append(entry)
                    elif d and d > TODAY: info["scheduled"].append(entry)
                    if entry["school"]:
                        out["schools_to_fellows"][entry["school"].lower()].append(f)
                    if entry["inv_status"]: info["inv_statuses"][entry["inv_status"]] += 1

        # Chain detection: for each prospect, check if "source" contains another prospect's first name
        prospect_firsts = {p["full_name_lower"].split()[0]: p["name"]
                           for p in info["prospects"] if p["full_name_lower"]}
        chain_map = defaultdict(list)
        for p in info["prospects"]:
            if not p["source"]: continue
            src_lower = p["source"].lower()
            # explicit "referred by X"
            m = re.search(r"referred by ([a-z][a-z ]+)", src_lower)
            hub_name = None
            if m: hub_name = m.group(1).strip().split()[0]
            else:
                for pfirst in prospect_firsts:
                    if pfirst and pfirst in src_lower and len(pfirst) > 2:
                        # avoid matching the prospect to themselves
                        if pfirst != p["full_name_lower"].split()[0]:
                            hub_name = pfirst
                            break
            if hub_name and hub_name in prospect_firsts:
                hub_initial = prospect_firsts[hub_name]
                chain_map[hub_initial].append(p["name"])
                info["source_buckets"]["Chain referral"] += 1
            else:
                info["source_buckets"][bucket_source(p["source"])] += 1
        info["chains"] = sorted(chain_map.items(), key=lambda x: -len(x[1]))
        info["chain_total"] = sum(len(refs) for _, refs in info["chains"])
        out["source_buckets"].update(info["source_buckets"])

        # de-dupe schools_to_fellows later
        info["districts"] = sorted(info["districts"])
        info["schools"] = sorted(info["schools"])

        # Derived counts
        info["leader_count"] = len(info["leaders"])
        info["ranked_leader_count"] = sum(1 for l in info["leaders"] if l["level"])
        info["prospect_count"] = len(info["prospects"])
        info["completed_count"] = len(info["completed"])
        info["scheduled_count"] = len(info["scheduled"])
        info["notes_count"] = sum(1 for c in info["completed"] if c["has_notes"])
        info["si_count"] = sum(1 for c in info["completed"] if c["has_si"])
        info["status"] = "active" if info["completed_count"] else "partial"

        out["fellows"][f] = info

    # Aggregate totals
    out["totals"] = {
        "fellows": len(fellows),
        "fellows_with_data": sum(1 for v in out["fellows"].values()
                                  if v["leader_count"] or v["prospect_count"]),
        "fellows_with_completed": sum(1 for v in out["fellows"].values() if v["completed_count"]),
        "leaders": sum(v["leader_count"] for v in out["fellows"].values()),
        "leaders_ranked": sum(v["ranked_leader_count"] for v in out["fellows"].values()),
        "prospects": sum(v["prospect_count"] for v in out["fellows"].values()),
        "completed": sum(v["completed_count"] for v in out["fellows"].values()),
        "scheduled": sum(v["scheduled_count"] for v in out["fellows"].values()),
        "with_notes": sum(v["notes_count"] for v in out["fellows"].values()),
        "with_si": sum(v["si_count"] for v in out["fellows"].values()),
    }
    # Cohort-overlap schools (where 2+ different fellows have prospects/leaders)
    overlap = {}
    for school, fellow_list in out["schools_to_fellows"].items():
        unique = sorted(set(fellow_list))
        if len(unique) >= 2:
            overlap[school] = unique
    out["overlap_schools"] = overlap
    out["schools_to_fellows"] = dict(out["schools_to_fellows"])
    out["source_buckets"] = dict(out["source_buckets"])

    # Voices: top 5 completed entries with notes AND si AND self_interest_text, newest first, max 1 per fellow
    voices = []
    seen_fellows = set()
    all_completed = []
    for fname, info in out["fellows"].items():
        for c in info["completed"]:
            if c["has_notes"] and c["has_si"] and c["self_interest_text"]:
                all_completed.append((fname, c))
    all_completed.sort(key=lambda x: x[1]["date"] or "", reverse=True)
    for fname, c in all_completed:
        if fname in seen_fellows: continue
        seen_fellows.add(fname)
        voices.append({
            "fellow": fname,
            "name": c["name"],
            "school": c["school"],
            "self_interest": c["self_interest_text"],
            "next_step": c["next_step_text"],
        })
        if len(voices) >= 5: break
    out["voices"] = voices

    # Three approaches (rule-based)
    # Pipeline: most prospects + scheduled.
    pipeline = max(out["fellows"].values(),
                   key=lambda v: v["prospect_count"] + v["scheduled_count"])
    # Depth: 100% notes rate (with ≥2 completed), tiebreak max completed.
    depth_candidates = [v for v in out["fellows"].values()
                        if v["completed_count"] >= 2
                        and v["notes_count"] == v["completed_count"]]
    depth = max(depth_candidates, key=lambda v: v["completed_count"]) if depth_candidates \
        else max(out["fellows"].values(),
                 key=lambda v: v["notes_count"] / max(1, v["completed_count"]))
    # Structure: most leaders identified (ranked or not).
    structure = max(out["fellows"].values(), key=lambda v: v["leader_count"])
    # Ensure all three are different fellows; if pipeline == depth (e.g., one dominant fellow),
    # pick the next-best for depth.
    chosen = {pipeline["name"]}
    if depth["name"] in chosen:
        for v in sorted(depth_candidates or out["fellows"].values(),
                        key=lambda v: -v["completed_count"]):
            if v["name"] not in chosen:
                depth = v; break
    chosen.add(depth["name"])
    if structure["name"] in chosen:
        for v in sorted(out["fellows"].values(), key=lambda v: -v["leader_count"]):
            if v["name"] not in chosen:
                structure = v; break
    out["approaches"] = {
        "pipeline": pipeline["name"],
        "depth": depth["name"],
        "structure": structure["name"],
    }

    return out

class _GspreadHeaderShim:
    """Mimics openpyxl ws.cell(row=1, column=c).value for header detection on gspread data."""
    def __init__(self, values):
        self.values = values
        self.max_column = max((len(r) for r in values), default=0)
    def cell(self, row, column):
        try: return self.values[row-1][column-1]
        except IndexError: return None

SOURCE_KEYWORDS = [
    ("Advocacy / org", ["mocsc", "naacp", "jwj", "jobs w", "equity group", "activism",
                         "moeep", "fhea", "wnea", "fhforward", "co-president", "director of",
                         "board member", "board pres", "board leader", "cohort", "fellowship",
                         "co-leader", "school board", "sccffps", "business contact", "advocacy"]),
    ("Event / forum", ["5/5 registration", "campaign launch", "campain launch", "foley",
                       "emergency meeting", "advocacy day", "forum", "labor event", "in person",
                       "in-person", "gym"]),
    ("Online",        ["online", "social media", "twitter", "facebook", "circles"]),
    ("Church / faith",["church", "faith", "congregation"]),
    ("Work",          ["coworker", "colleague"]),
    ("School / PTA",  ["pta", "pto", "classroom", "school dance", "fellow ", "parent",
                       "teacher", "principal", "volunteering @ school", "classmate",
                       "school activity", "fhcentral", "school"]),
    ("Friend / family", ["friend", "family", "neighbor", "gpa", "sister", "cousin"]),
]
def bucket_source(s):
    sl = (s or "").lower()
    if not sl.strip(): return "Other"
    for label, kws in SOURCE_KEYWORDS:
        if any(k in sl for k in kws): return label
    return "Other"

# ── RENDERING ───────────────────────────────────────────────────────────────

def compute_deltas(curr, prev):
    """Return dict of delta strings keyed by metric name."""
    if not prev: return {}
    deltas = {}
    for k in ("fellows", "fellows_with_data", "leaders", "prospects", "completed", "scheduled"):
        c, p = curr["totals"].get(k, 0), prev.get("totals", {}).get(k, 0)
        diff = c - p
        sign = "+" if diff > 0 else ""
        deltas[k] = f"{sign}{diff}"
    return deltas

def resample_to_sundays(history, today):
    """Resample history to weekly-Sunday snapshots. For each Sunday between the
    earliest snapshot and today, use the latest snapshot whose date ≤ that Sunday.
    Append today's snapshot as a final point if it isn't already a Sunday."""
    if not history: return []
    from datetime import timedelta, date as date_cls
    def parse(s): return date_cls.fromisoformat(s) if isinstance(s, str) else s
    snap_dates = [(parse(h["date"]), h) for h in history]
    snap_dates.sort(key=lambda x: x[0])
    first = snap_dates[0][0]
    # Find first Sunday on or after `first`
    sunday = first + timedelta(days=(6 - first.weekday()) % 7)
    points = []
    while sunday <= today:
        # Use latest snapshot with date <= sunday
        usable = [(d, h) for d, h in snap_dates if d <= sunday]
        if usable:
            d, h = usable[-1]
            points.append({**h, "label_date": sunday.isoformat()})
        sunday += timedelta(days=7)
    # Append today as final point if today != last Sunday
    last_sunday = sunday - timedelta(days=7)
    if last_sunday != today and snap_dates and snap_dates[-1][0] == today:
        points.append({**snap_dates[-1][1], "label_date": today.isoformat()})
    return points

def find_week_ago_snapshot(history, today):
    """Return the snapshot from ~7 days ago for week-over-week deltas."""
    from datetime import timedelta, date as date_cls
    if not history: return None
    target = today - timedelta(days=7)
    def parse(s): return date_cls.fromisoformat(s) if isinstance(s, str) else s
    older = [h for h in history if parse(h["date"]) <= target]
    return older[-1] if older else None

NETWORK_LEAD_INS = {
    "Chain referral": "the organizing flywheel is the dominant motion — most new prospects come through people the cohort has already met.",
    "School / PTA": "the cohort is rooted in school communities — most prospects come from parents and teachers in the buildings fellows are working.",
    "Advocacy / org": "the cohort is plugged into organized advocacy spaces — most prospects come through existing orgs and coalitions.",
    "Online": "online channels lead — broad reach, less personal warmth. Watch chain referrals to see relationships deepen.",
    "Event / forum": "events are the dominant entry point — single-shot moments that need follow-up to turn into relationships.",
    "Friend / family / neighbor": "personal networks are the dominant source — strong starting trust, will need to expand beyond.",
    "Friend / family": "personal networks are the dominant source — strong starting trust, will need to expand beyond.",
    "Work": "work relationships lead the cohort's prospect base.",
    "Church / faith": "faith communities are the dominant source.",
    "Other": "sources are varied — no single channel dominates.",
}

def generate_reflection_questions(data, history, today):
    """Generate dated, data-driven reflection questions."""
    from datetime import timedelta, date as date_cls
    # Monday of this week
    monday = today - timedelta(days=today.weekday())
    week_ago = find_week_ago_snapshot(history, today)
    fellows = data["fellows"]

    # Q2: gap-driven — fellows still at zero completed
    zero_completed = sorted([v["name"] for v in fellows.values()
                              if v["completed_count"] == 0])
    # Q3: bright-spot — top performer by completed count this week
    if week_ago:
        # We don't have per-fellow last-week data; use current totals as proxy for "this week"
        pass
    top = max(fellows.values(), key=lambda v: v["completed_count"]) if fellows else None
    top_quality = max(fellows.values(),
                      key=lambda v: (v["notes_count"], v["si_count"])) if fellows else None
    # Q for notes capture
    no_notes = [v["name"] for v in fellows.values()
                 if v["completed_count"] and v["notes_count"] < v["completed_count"]]

    questions = []

    # Q1: always thematic
    questions.append(
        "Look at someone whose self-interest you wrote down. What did the act of writing it down teach you about the conversation you actually had?")

    # Q2: data-driven on coverage gap (or notes gap if everyone's started)
    if zero_completed:
        names = ", ".join(zero_completed[:3]) + ("…" if len(zero_completed) > 3 else "")
        questions.append(
            f"{len(zero_completed)} fellow{'s' if len(zero_completed)!=1 else ''} still at zero 1-on-1s ({names}). What's between you and your first conversation?")
    elif no_notes:
        questions.append(
            f"{len(no_notes)} fellow{'s' if len(no_notes)!=1 else ''} have completed 1-on-1s without notes. What stops you from getting to the spreadsheet after the meeting?")
    else:
        questions.append(
            "Every fellow has logged at least one 1-on-1 with notes. What's the next conversation you've been avoiding?")

    # Q3: data-driven bright spot
    if top and top["completed_count"] >= 3:
        questions.append(
            f"{top['name']} has logged {top['completed_count']} 1-on-1s. What did they figure out that we can borrow?")
    else:
        questions.append(
            "Look at the names on your list. Which of them could you invite to take a specific action this week?")

    # Q4: constant practice prompt
    questions.append(
        "What's one 1-on-1 you could do between now and next Monday? Name the person.")

    return {
        "week_of": monday.strftime("%B %-d, %Y"),
        "questions": questions,
    }

def render(data, deltas, history):
    from jinja2 import Environment, FileSystemLoader, select_autoescape
    env = Environment(loader=FileSystemLoader(str(ROOT / "scripts")),
                      autoescape=select_autoescape(["html"]))
    env.filters["initial"] = initial
    tpl = env.get_template("template.html")

    # Week-over-week deltas
    week_ago_snap = find_week_ago_snapshot(history, TODAY)
    wow = {}
    for k in ("fellows_with_data", "leaders", "prospects", "completed", "scheduled", "with_notes"):
        curr_v = data["totals"].get(k, 0)
        prev_v = (week_ago_snap or {}).get(k, 0)
        diff = curr_v - prev_v
        if prev_v == 0:
            pct = None
        else:
            pct = int(round((diff / prev_v) * 100))
        wow[k] = {"current": curr_v, "prev": prev_v, "diff": diff, "pct": pct}

    # Resample history to weekly-Sunday points for the chart
    weekly = resample_to_sundays(history, TODAY)

    # Build chart panels from weekly snapshots
    chart_w, chart_h = 700, 220
    margin_l, margin_r, margin_t, margin_b = 36, 28, 28, 52  # more bottom for date labels
    plot_w = chart_w - margin_l - margin_r
    plot_h = chart_h - margin_t - margin_b
    series_meta = [
        ("completed", "Completed 1-on-1s", "#22c55e"),
        ("prospects", "Prospects logged", "#8b5cf6"),
        ("leaders", "Leaders identified", "#5371ff"),
        ("fellows_with_data", "Fellows w/ data entry", "#f59e0b"),
    ]
    panels = []
    if weekly:
        n = len(weekly)
        x_step = plot_w / max(1, n - 1) if n > 1 else 0
        for key, label, color in series_meta:
            vals = [int(h.get(key, 0) or 0) for h in weekly]
            max_v = max(vals) if vals else 1
            if max_v == 0: max_v = 1
            points = []
            for i, v in enumerate(vals):
                x = margin_l + (i * x_step if n > 1 else plot_w / 2)
                y = margin_t + plot_h - (v / max_v) * plot_h
                d = weekly[i]["label_date"]
                # Format date as M/D
                mo, dd = d.split("-")[1].lstrip("0"), d.split("-")[2].lstrip("0")
                points.append({"x": round(x, 1), "y": round(y, 1),
                               "value": v, "date": d, "label": f"{mo}/{dd}"})
            wow_diff = vals[-1] - (week_ago_snap or {}).get(key, 0) if week_ago_snap else None
            panels.append({
                "key": key, "label": label, "color": color,
                "points": points, "max": max_v,
                "current": vals[-1], "start": vals[0],
                "delta": vals[-1] - vals[0],
                "wow_diff": wow_diff,
            })

    chart = {
        "w": chart_w, "h": chart_h,
        "margin": {"l": margin_l, "r": margin_r, "t": margin_t, "b": margin_b},
        "plot_w": plot_w, "plot_h": plot_h,
        "panels": panels,
        "history": history,
        "weekly": weekly,
    }

    # Sort fellows for grid: completed desc, prospects desc
    fellows_sorted = sorted(data["fellows"].values(),
                            key=lambda v: (-v["completed_count"], -v["prospect_count"], -v["leader_count"]))
    # For bar chart: prospects desc
    bar_fellows = sorted(data["fellows"].values(),
                          key=lambda v: (-v["prospect_count"], -v["leader_count"]))
    max_prospects = max((v["prospect_count"] for v in data["fellows"].values()), default=1) or 1

    # Quality table: completed desc, then notes pct
    quality = sorted(data["fellows"].values(),
                     key=lambda v: (-v["completed_count"],
                                    -(v["notes_count"]/max(1,v["completed_count"])),
                                    -v["prospect_count"]))

    # Leaders slide cards: ranked leaders desc, show top 6 with ≥1 ranked leader
    leader_cards = [v for v in sorted(data["fellows"].values(),
                                       key=lambda x: -x["ranked_leader_count"])
                    if v["ranked_leader_count"] > 0][:6]
    leaders_started = sum(1 for v in data["fellows"].values() if v["ranked_leader_count"] > 0)
    no_leader_list = [v["name"] for v in data["fellows"].values() if v["ranked_leader_count"] == 0]

    # Approaches
    apps = data["approaches"]
    approach_fellows = {k: data["fellows"][apps[k]] for k in apps if apps.get(k)}

    # Chain cards: top 3 fellows by chain referral count
    fellow_chain_counts = [(v, sum(len(refs) for _, refs in v["chains"])) for v in data["fellows"].values()]
    fellow_chain_counts.sort(key=lambda x: -x[1])
    chain_cards = [v for v, count in fellow_chain_counts if count > 0][:3]

    # Source buckets for breakdown
    total_sources = sum(data["source_buckets"].values()) or 1
    buckets = sorted(data["source_buckets"].items(), key=lambda x: -x[1])

    # Dynamic network-slide commentary: lead-in adapts to which bucket is on top
    if buckets:
        top_label, top_count = buckets[0]
        top_pct = int(round(top_count / total_sources * 100))
        lead_in = NETWORK_LEAD_INS.get(top_label, f"{top_label} leads as the dominant source.")
        network_commentary = {
            "top_label": top_label,
            "top_pct": top_pct,
            "top_count": top_count,
            "lead_in": lead_in,
        }
    else:
        network_commentary = {"top_label": "", "top_pct": 0, "top_count": 0, "lead_in": ""}

    # Reflection questions (dynamic)
    reflection = generate_reflection_questions(data, history, TODAY)

    return tpl.render(
        d=data,
        deltas=deltas,
        today=TODAY,
        built_at=dt.datetime.now().strftime("%b %-d, %Y at %-I:%M %p"),
        fellows_sorted=fellows_sorted,
        bar_fellows=bar_fellows,
        max_prospects=max_prospects,
        quality=quality,
        leader_cards=leader_cards,
        leaders_started=leaders_started,
        no_leader_list=no_leader_list,
        approach_fellows=approach_fellows,
        chain_cards=chain_cards,
        buckets=buckets,
        total_sources=total_sources,
        chart=chart,
        wow=wow,
        network_commentary=network_commentary,
        reflection=reflection,
    )

# ── MAIN ────────────────────────────────────────────────────────────────────

def load_history():
    """Return list of {date, fellows_with_data, completed, prospects, leaders, ...} snapshots."""
    history_path = DATA_DIR / "history.jsonl"
    if not history_path.exists(): return []
    out = []
    for line in history_path.read_text().splitlines():
        if line.strip():
            out.append(json.loads(line))
    out.sort(key=lambda h: h.get("date", ""))
    return out

def save_history(data):
    history_path = DATA_DIR / "history.jsonl"
    history = load_history()
    snapshot = {"date": TODAY.isoformat(), **data["totals"]}
    history = [h for h in history if h.get("date") != snapshot["date"]]
    history.append(snapshot)
    history.sort(key=lambda h: h["date"])
    history_path.write_text("\n".join(json.dumps(h) for h in history) + "\n")
    print(f"→ Wrote {history_path} ({len(history)} snapshots)", file=sys.stderr)
    return history

def main():
    print(f"Building POF dashboard for {TODAY}", file=sys.stderr)
    workbook = load_workbook_sheets()
    if workbook is None:
        print("→ Using local xlsx (no GOOGLE_SERVICE_ACCOUNT_JSON set)", file=sys.stderr)
        workbook = load_workbook_local()
    else:
        print("→ Using Google Sheet", file=sys.stderr)

    data = aggregate(workbook)
    print(f"   Fellows: {data['totals']['fellows']}", file=sys.stderr)
    print(f"   Completed: {data['totals']['completed']}", file=sys.stderr)
    print(f"   Scheduled: {data['totals']['scheduled']}", file=sys.stderr)
    print(f"   Warnings: {data['warnings'] or 'none'}", file=sys.stderr)

    # Load prior build for deltas
    prev = None
    prev_path = DATA_DIR / "data-previous.json"
    if prev_path.exists():
        prev = json.loads(prev_path.read_text())
    deltas = compute_deltas(data, prev)

    # Persist data + append snapshot to history BEFORE rendering so the chart includes today.
    DATA_DIR.mkdir(exist_ok=True)
    curr_path = DATA_DIR / "data.json"
    if curr_path.exists():
        prev_path.write_text(curr_path.read_text())
    curr_path.write_text(json.dumps(data, indent=2, default=str))
    print(f"→ Wrote {curr_path}", file=sys.stderr)
    history = save_history(data)

    # Render
    html = render(data, deltas, history)
    OUTPUT_HTML.write_text(html)
    print(f"→ Wrote {OUTPUT_HTML}", file=sys.stderr)

if __name__ == "__main__":
    main()
